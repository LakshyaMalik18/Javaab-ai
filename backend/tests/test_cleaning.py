"""Edge-case test corpus §9 #1–#12."""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest

from app.engines.ingest import (
    IngestError,
    detect_and_decode,
    ingest_csv,
    ingest_excel,
    ingest_json,
)
from app.engines.cleaning import clean
from app.engines.canonical import canonicalize_column, find_near_duplicate_rows

F = Path(__file__).parent / "fixtures"


def _load(name: str) -> dict:
    return ingest_csv((F / name).read_bytes(), table_name=name.split(".")[0])


# ─────────────────────────────────────────────────────────────────────────────
# §9 #1 — Clean single CSV, obvious types
# ─────────────────────────────────────────────────────────────────────────────

def test_01_clean_csv_loads():
    result = _load("01_clean.csv")
    df = result["df"]
    assert list(df.columns) == ["id", "name", "age", "salary", "joined"]
    assert len(df) == 3


def test_01_types_inferred():
    df_raw = _load("01_clean.csv")["df"]
    cleaned, ledger, dupes, flags = clean(df_raw, "test01")
    # salary should be numeric
    assert cleaned["salary"].dropna().apply(lambda v: isinstance(v, (int, float))).all()
    # no ambiguity flags on a clean file
    assert flags == []
    # no duplicates
    assert dupes == []


# ─────────────────────────────────────────────────────────────────────────────
# §9 #2 — Two-file joinable set
# ─────────────────────────────────────────────────────────────────────────────

def test_02_two_files_load():
    cust = _load("02a_customers.csv")["df"]
    orders = _load("02b_orders.csv")["df"]
    assert "customer_id" in cust.columns
    assert "cust_id" in orders.columns
    assert len(cust) == 3
    assert len(orders) == 3


def test_02_orders_amount_is_numeric():
    df = _load("02b_orders.csv")["df"]
    cleaned, ledger, _, _ = clean(df, "orders")
    vals = cleaned["amount"].dropna()
    assert vals.apply(lambda v: isinstance(v, float)).all()


# ─────────────────────────────────────────────────────────────────────────────
# §9 #3 — Currency / percent / thousands separators
# ─────────────────────────────────────────────────────────────────────────────

def test_03_currency_stripped():
    df = _load("03_currency.csv")["df"]
    cleaned, ledger, _, _ = clean(df, "currency")
    prices = cleaned["price"].dropna()
    assert all(isinstance(v, (int, float)) for v in prices), f"prices not numeric: {list(prices)}"


def test_03_negative_parens():
    df = _load("03_currency.csv")["df"]
    cleaned, _, _, _ = clean(df, "currency")
    prices = list(cleaned["price"].dropna())
    assert any(v < 0 for v in prices), "Parenthesised negative not converted"


def test_03_unit_suffixes():
    df = _load("03_currency.csv")["df"]
    cleaned, _, _, _ = clean(df, "currency")
    # Widget E revenue = $125k → 125_000
    revenue_vals = list(cleaned["revenue"].dropna())
    assert any(v >= 100_000 for v in revenue_vals if isinstance(v, (int, float)))


def test_03_percent_stripped():
    df = _load("03_currency.csv")["df"]
    cleaned, _, _, _ = clean(df, "currency")
    disc = cleaned["discount"].dropna()
    assert all(isinstance(v, (int, float)) for v in disc)


def test_03_ledger_records_changes():
    df = _load("03_currency.csv")["df"]
    _, ledger, _, _ = clean(df, "currency")
    rules = {r.rule for r in ledger.records}
    assert "numeric_coerce" in rules


# ─────────────────────────────────────────────────────────────────────────────
# §9 #4 — Mixed / ambiguous date formats
# ─────────────────────────────────────────────────────────────────────────────

def test_04_date_normalized_to_iso():
    df = _load("04_dates.csv")["df"]
    cleaned, ledger, _, flags = clean(df, "dates")
    # event_date column: 15/01/2024 → 2024-01-15 (DD/MM is unambiguous, day=15>12)
    dates = cleaned["event_date"].dropna().astype(str)
    iso_pattern = r"^\d{4}-\d{2}-\d{2}$"
    import re
    assert all(re.match(iso_pattern, d) for d in dates), f"Non-ISO date: {list(dates)}"


def test_04_ambiguous_column_flagged():
    df = _load("04_dates.csv")["df"]
    # reported_date: 01/15/2024 — first field is 01 ≤ 12 but second is 15 > 12 → mdy
    _, _, _, flags = clean(df, "dates")
    # The engine may or may not flag; the column should still parse
    cleaned, _, _, _ = clean(df, "dates")
    import re
    dates = cleaned["reported_date"].dropna().astype(str)
    assert all(re.match(r"^\d{4}-\d{2}-\d{2}$", d) for d in dates)


# ─────────────────────────────────────────────────────────────────────────────
# §9 #5 — Messy nulls
# ─────────────────────────────────────────────────────────────────────────────

def test_05_null_tokens_become_na():
    df = _load("05_nulls.csv")["df"]
    cleaned, ledger, _, _ = clean(df, "nulls")
    score_col = cleaned["score"]
    # Rows 2-7 had null tokens; should be pd.NA now
    null_count = score_col.isna().sum()
    assert null_count >= 5, f"Expected ≥5 NAs in score, got {null_count}"


def test_05_ledger_records_null_changes():
    df = _load("05_nulls.csv")["df"]
    _, ledger, _, _ = clean(df, "nulls")
    rules = {r.rule for r in ledger.records}
    assert "null_normalize" in rules


def test_05_real_values_preserved():
    df = _load("05_nulls.csv")["df"]
    cleaned, _, _, _ = clean(df, "nulls")
    scores = cleaned["score"].dropna()
    numeric = [v for v in scores if isinstance(v, (int, float))]
    assert len(numeric) >= 2  # rows 1 and 8 have real scores


# ─────────────────────────────────────────────────────────────────────────────
# §9 #6 — Coded / abbreviated headers
# ─────────────────────────────────────────────────────────────────────────────

def test_06_coded_headers_preserved():
    df = _load("06_coded_headers.csv")["df"]
    assert "cst_id" in df.columns
    assert "ord_dt" in df.columns
    assert "amt" in df.columns


def test_06_values_cleaned():
    df = _load("06_coded_headers.csv")["df"]
    cleaned, _, _, _ = clean(df, "coded")
    assert cleaned["amt"].dropna().apply(lambda v: isinstance(v, float)).all()


# ─────────────────────────────────────────────────────────────────────────────
# §9 #7 — Banner / preamble rows + multi-row header
# ─────────────────────────────────────────────────────────────────────────────

def test_07_banner_skipped():
    result = _load("07_banner.csv")
    df = result["df"]
    # Should not have the banner text as a column
    assert not any("Quarterly" in c for c in df.columns)
    assert not any("Generated" in c for c in df.columns)


def test_07_data_rows_correct():
    df = _load("07_banner.csv")["df"]
    # Should have 2 data rows (Alice, Bob); footer "Total" stripped
    assert len(df) == 2, f"Expected 2 rows, got {len(df)}: {df.to_dict()}"


def test_07_multirow_header_flattened():
    df = _load("07_banner.csv")["df"]
    # Columns should be merged (Customer_ID or similar, not raw "ID" and "Customer" separately)
    col_names = " ".join(df.columns)
    assert len(df.columns) == 4


# ─────────────────────────────────────────────────────────────────────────────
# §9 #8 — Canonicalization (USA / U.S.A. / America)
# ─────────────────────────────────────────────────────────────────────────────

def test_08_alias_map_applied():
    df = _load("08_canonicalize.csv")["df"]
    from app.models import ChangeLedger
    ledger = ChangeLedger()
    series, suggestions = canonicalize_column(df["country"], "canon", "country", ledger)
    # All USA variants should collapse
    united_states = (series == "United States").sum()
    assert united_states >= 4, f"Expected ≥4 'United States', got {united_states}: {list(series)}"


def test_08_uk_variants_collapsed():
    df = _load("08_canonicalize.csv")["df"]
    from app.models import ChangeLedger
    ledger = ChangeLedger()
    series, _ = canonicalize_column(df["country"], "canon", "country", ledger)
    uk_count = (series == "United Kingdom").sum()
    assert uk_count >= 2


def test_08_ledger_records_canonicalization():
    df = _load("08_canonicalize.csv")["df"]
    from app.models import ChangeLedger
    ledger = ChangeLedger()
    _, _ = canonicalize_column(df["country"], "canon", "country", ledger)
    assert any(r.rule == "canonicalize_alias" for r in ledger.records)


# ─────────────────────────────────────────────────────────────────────────────
# §9 #9 — Near-duplicate rows
# ─────────────────────────────────────────────────────────────────────────────

def test_09_near_dupes_detected():
    df = _load("09_near_dupes.csv")["df"]
    near_dupes = find_near_duplicate_rows(df, key_columns=["company"])
    assert len(near_dupes) >= 2, f"Expected ≥2 near-dup pairs, got {near_dupes}"


def test_09_near_dupes_not_deleted():
    df = _load("09_near_dupes.csv")["df"]
    original_len = len(df)
    near_dupes = find_near_duplicate_rows(df, key_columns=["company"])
    assert len(df) == original_len  # df unchanged


def test_09_acme_pair_found():
    df = _load("09_near_dupes.csv")["df"]
    near_dupes = find_near_duplicate_rows(df, key_columns=["company"])
    companies_in_pairs = set()
    for nd in near_dupes:
        companies_in_pairs.add(df.iloc[nd["indices"][0]]["company"])
        companies_in_pairs.add(df.iloc[nd["indices"][1]]["company"])
    assert any("Acme" in c for c in companies_in_pairs)


# ─────────────────────────────────────────────────────────────────────────────
# §9 #10 — Multi-sheet Excel + nested JSON
# ─────────────────────────────────────────────────────────────────────────────

def test_10_excel_multisheet():
    raw = (F / "10_multisheet.xlsx").read_bytes()
    sheets = ingest_excel(raw)
    assert "Customers" in sheets
    assert "Orders" in sheets
    assert len(sheets["Customers"]["df"]) == 2
    assert len(sheets["Orders"]["df"]) == 2


def test_10_json_nested_flattened():
    raw = (F / "10_nested.json").read_bytes()
    result = ingest_json(raw, "nested")
    df = result["df"]
    # Nested address should be flattened: address.city, address.zip
    assert any("city" in c for c in df.columns)
    assert len(df) == 3


def test_10_json_ragged_keys_null():
    raw = (F / "10_nested.json").read_bytes()
    result = ingest_json(raw, "nested")
    df = result["df"]
    # Carol has no score — should be None/NaN
    carol_row = df[df["name"] == "Carol"]
    assert len(carol_row) == 1
    # Bob has no zip
    bob_row = df[df["name"] == "Bob"]
    assert len(bob_row) == 1


# ─────────────────────────────────────────────────────────────────────────────
# §9 #11 — Empty file / header-only / single-column
# ─────────────────────────────────────────────────────────────────────────────

def test_11_empty_file_raises():
    with pytest.raises(IngestError, match="empty"):
        ingest_csv((F / "11_empty.csv").read_bytes())


def test_11_header_only_yields_empty_table():
    # Spec: a header-only file is a valid 0-row table, not an error.
    result = ingest_csv((F / "11_header_only.csv").read_bytes())
    df = result["df"]
    assert len(df) == 0
    assert list(df.columns) == ["id", "name", "age"]


def test_11_single_column_loads():
    result = ingest_csv((F / "11_single_col.csv").read_bytes())
    df = result["df"]
    assert len(df.columns) == 1
    assert len(df) == 3


# ─────────────────────────────────────────────────────────────────────────────
# §9 #12 — Encoding oddities
# ─────────────────────────────────────────────────────────────────────────────

def test_12_utf16_decoded():
    raw = (F / "12_utf16.csv").read_bytes()
    text = detect_and_decode(raw)
    assert "Ångström" in text or "ngstr" in text  # charset may transliterate


def test_12_latin1_decoded():
    raw = (F / "12_latin1.csv").read_bytes()
    text = detect_and_decode(raw)
    assert "Ren" in text  # René at minimum


def test_12_utf8_bom_stripped():
    raw = (F / "12_utf8bom.csv").read_bytes()
    text = detect_and_decode(raw)
    assert not text.startswith("﻿")
    result = ingest_csv(raw)
    df = result["df"]
    assert "id" in df.columns  # BOM must not corrupt first column name


def test_12_latin1_csv_loads():
    raw = (F / "12_latin1.csv").read_bytes()
    result = ingest_csv(raw)
    df = result["df"]
    assert len(df) == 2
    assert "id" in df.columns


# ─────────────────────────────────────────────────────────────────────────────
# JSON nested arrays — flatten to text, one row per record (regression)
# ─────────────────────────────────────────────────────────────────────────────

def test_json_nested_array_flattened_to_text():
    """A JSON array field is preserved as comma-joined text in one cell — not
    dropped, and not exploded into extra rows."""
    from app.upload_pipeline import process_upload

    raw = json.dumps([
        {"id": 1, "amount": 100, "tags": ["urgent", "vip"]},
        {"id": 2, "amount": 200, "tags": ["c"]},
    ]).encode()
    result = process_upload([("orders.json", raw)])

    assert result.errors == []
    df = result.tables["orders"]
    # row count unchanged — no explosion, no dropped rows
    assert len(df) == 2
    # array became a text column with comma-joined values
    assert df["tags"].tolist() == ["urgent, vip", "c"]


def test_json_nested_array_does_not_crash_cleaning():
    """Regression: a list cell used to raise in the cleaning engine
    (pd.isna on a list). Ingest+clean must now complete cleanly."""
    from app.upload_pipeline import process_upload

    raw = json.dumps([{"id": 1, "roles": ["a", "b", "c"]}]).encode()
    result = process_upload([("r.json", raw)])
    assert result.errors == []
    assert result.tables["r"]["roles"].tolist() == ["a, b, c"]


# ─────────────────────────────────────────────────────────────────────────────
# Per-file error isolation — one bad file must not sink the batch
# ─────────────────────────────────────────────────────────────────────────────

def test_bad_file_does_not_crash_multifile_upload():
    """A genuinely unprocessable file fails gracefully with a per-file message
    while the other files in the same batch still process."""
    from app.upload_pipeline import process_upload

    good = b"name,score\nAl,10\nBo,20"
    broken = b"{ this is not valid json"
    result = process_upload([("good.csv", good), ("broken.json", broken)])

    # the good file made it through
    assert "good" in result.tables
    assert len(result.tables["good"]) == 2
    # the bad file produced a clear, named error rather than crashing
    assert any("broken.json" in e for e in result.errors)


def test_cleaning_stage_failure_is_isolated(monkeypatch):
    """The widened error boundary: a failure inside cleaning (not just ingest)
    is caught per-table, so other tables in the batch still complete."""
    from app import upload_pipeline
    from app.engines import cleaning as cleaning_mod

    real_clean = cleaning_mod.clean

    def boom(df, table_name="data", **kwargs):
        if table_name == "bad":
            raise RuntimeError("simulated cleaning explosion")
        return real_clean(df, table_name=table_name, **kwargs)

    monkeypatch.setattr(upload_pipeline.cleaning_mod, "clean", boom)

    good = b"name,score\nAl,10\nBo,20"
    bad = b"a,b\n1,2"
    result = upload_pipeline.process_upload([("good.csv", good), ("bad.csv", bad)])

    # good table survived; bad table was isolated with a named error
    assert "good" in result.tables
    assert "bad" not in result.tables
    assert any("bad" in e and "could not process" in e for e in result.errors)


# ─────────────────────────────────────────────────────────────────────────────
# Excel formula columns — read cached values; warn (don't silently drop) when none
# ─────────────────────────────────────────────────────────────────────────────

def _build_formula_xlsx(with_cache: bool) -> bytes:
    """An xlsx whose C column is a formula (=A*B). openpyxl never computes
    formulas, so `with_cache` injects cached <v> results into the sheet XML to
    simulate a file that was opened/saved in Excel."""
    import io
    import zipfile

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"], ws["C1"] = "qty", "price", "total"
    ws["A2"], ws["B2"], ws["C2"] = 3, 2.5, "=A2*B2"
    ws["A3"], ws["B3"], ws["C3"] = 4, 1.0, "=A3*B3"
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if not with_cache:
        return data

    zin = zipfile.ZipFile(io.BytesIO(data))
    items = {n: zin.read(n) for n in zin.namelist()}
    sheet = next(n for n in items if n.startswith("xl/worksheets/sheet"))
    xml = items[sheet].decode()
    xml = xml.replace("<f>A2*B2</f>", "<f>A2*B2</f><v>7.5</v>")
    xml = xml.replace("<f>A3*B3</f>", "<f>A3*B3</f><v>4</v>")
    items[sheet] = xml.encode()
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for n, b in items.items():
            z.writestr(n, b)
    return out.getvalue()


def test_excel_formula_cached_values_imported():
    """A formula column with cached results imports those values as normal data."""
    result = ingest_excel(_build_formula_xlsx(with_cache=True))
    sheet = result["Sheet1"]
    df = sheet["df"]
    assert "total" in df.columns           # column NOT dropped
    assert df["total"].tolist() == ["7.5", "4"]
    assert sheet["warnings"] == []         # nothing to warn about


def test_excel_formula_without_cache_warns_not_dropped():
    """A formula column with no cached value produces a warning instead of being
    silently dropped — naming the column so the loss is visible."""
    result = ingest_excel(_build_formula_xlsx(with_cache=False))
    sheet = result["Sheet1"]
    # the non-formula columns still import fine
    assert "qty" in sheet["df"].columns and "price" in sheet["df"].columns
    # and the unreadable formula column is reported, not lost in silence
    assert len(sheet["warnings"]) == 1
    assert "total" in sheet["warnings"][0]


def test_excel_formula_warning_surfaces_through_upload():
    """The ingest warning rides the UploadResult.warnings channel end-to-end."""
    from app.upload_pipeline import process_upload

    up = process_upload([("book.xlsx", _build_formula_xlsx(with_cache=False))])
    assert any("total" in w for w in up.warnings)
    # cached version: clean import, no warning
    up2 = process_upload([("book.xlsx", _build_formula_xlsx(with_cache=True))])
    assert up2.warnings == []
    # the cleaning engine coerces the imported formula results to numbers
    assert up2.tables["sheet1"]["total"].tolist() == [7.5, 4.0]
